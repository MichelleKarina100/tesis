import json

from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView,TemplateView, View
from django.db import transaction
from django.db.models import Q
from datetime import datetime
from django.utils import timezone

from core.matricula.forms import *
from core.security.mixins import PermissionMixin
from core.matricula.models import * 
from core.homepage.models import *
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http import JsonResponse, HttpResponse,HttpResponseRedirect
from django.template.loader import get_template
from weasyprint import HTML, CSS

from core.cursos.models import *
from crum import get_current_user
from openpyxl import load_workbook
from django.db import transaction
class VerCursoListView(CreateView):
    template_name = 'matricula/verCurso.html'
    form_class = MatriculaForm

    def get_success_url(self):
        pk = self.kwargs['pk']
        return reverse_lazy('verCurso', kwargs={'pk': pk})
  
    def validate_data(self):
            data = {'valid': True}
            try:
                type = self.request.POST['type']
                obj = self.request.POST['obj'].strip()
                obj2 = self.request.POST['obj2'].strip()
                obj3 = 'Evaluacion'
                pk = self.kwargs['pk']

                if type == 'estudiante':
                    if Matricula.objects.filter(estudiante=obj,curso_id=pk):
                        data['valid'] = False
                elif type == 'name':
                    if CrearTarea.objects.filter(name__icontains=obj,curso_id=pk):
                        data['valid'] = False
                elif type == 'producto':
                    if Silabo.objects.filter(quimestre__icontains=obj,producto__icontains=obj2,curso_id=pk):
                        data['valid'] = False
            except:
                pass
            return JsonResponse(data)

    def post(self, request, *args, **kwargs):
        action = request.POST['action']
        data = {}
        try:
            if action == 'add':
                with transaction.atomic():
                    print('Accion Agregar Matricula',action)
                    estudiante = request.POST['estudiante']
                    print('estudiante',estudiante)
                    observacion = request.POST['observacion']
                    print('observacion',observacion)
                    curso = self.kwargs['pk']
                    e = Matricula()
                    e.estudiante_id = estudiante
                    e.curso_id = curso
                    e.profesor_id = request.user.id
                    e.observacion = observacion
                    e.save()    
            elif action == 'post_add':
                with transaction.atomic():
                    print('Accion Agregar Foro',action)
                    message = request.POST['publicacion']
                    print('mensaje de la publicacion: ',message)
                    curso_id = self.kwargs['pk']
                    pb = Post()
                    pb.publicacion = message
                    pb.curso_id = curso_id
                    pb.user_id = request.user.id
                    pb.save()
            elif action == 'add_list':
                with transaction.atomic():
                    check = request.POST.getlist('checks[]')                    
                    for i in check:
                        lista = ListaEstudiantes ()
                        lista.estudiant_id = i
                        lista.curso_id = self.kwargs['pk']
                        lista.asistencia = False
                        lista.save()

            elif action == 'add_homework':
                with transaction.atomic():
                    
                    
                    id_curso = self.kwargs['pk']
                    estudiants = Matricula.objects.filter(curso=id_curso)
                    tema = request.POST['name']
                    #print('tema: ',tema)
                    juego = request.POST['silabo']
                    #print('juego: ',juego)
                    descripcion = request.POST['descripcion']
                    #print('descripcion: ', descripcion)
                    fecha = request.POST['creada_en']
                    #producto = request.POST.get('producto')

                    crear = CrearTarea()
                    crear.name = tema
                    crear.curso_id = id_curso
                    crear.silabo_id = juego
                    crear.descripcion = descripcion
                    crear.creada_en = fecha
                    crear.save()
                    tareaActual = CrearTarea.objects.get(name__icontains = tema)
                    #print('Tarea Actual: ',tareaActual.id)
                    if tareaActual :               
                        for i in estudiants:
                            print('Tarea Actual: ', tareaActual.id)
                            print('estudiante: ', i.estudiante_id)
                            entregar = EntregarTarea()
                            entregar.tarea_id = tareaActual.id
                            entregar.creada_en = fecha
                            entregar.estudiante_id = i.estudiante_id
                            entregar.save()
            
            elif action == 'add_silabo':
                with transaction.atomic():
                    print('seccion silabo')
                    id_claseDemo = request.POST['id_claseDemo']
                    print('este valor ',id_claseDemo)
                    form = SilaboForm(request.POST, files=request.FILES)
                    if form.is_valid():
                        pk = self.kwargs['pk']
                        form.instance.curso_id = pk
                        form.instance.claseDemo = id_claseDemo
                        form.save()
                        print("guardar")
                    else:
                        print("nell")
            elif action == 'upload_excel':
                print("WWWWWWWWW")
                with transaction.atomic():
                    archive = request.FILES['archive']
                    workbook = load_workbook(filename=archive, data_only=True)
                    excel = workbook[workbook.sheetnames[0]]
                    for row in range(2, excel.max_row + 1):

                        #id = int(excel.cell(row=row, column=1).value)
                        actividad = excel.cell(row=row, column=1).value
                        descripcion = excel.cell(row=row, column=2).value
                        quimestre = excel.cell(row=row, column=3).value
                        producto = excel.cell(row=row, column=4).value
                        url = excel.cell(row=row, column=5).value

                     
                        a = Silabo()
                        a.actividad=actividad
                        a.descripcion= descripcion
                        a.quimestre= quimestre
                        a.producto= producto
                        a.url= url
                        a.curso_id = self.kwargs['pk']
                        a.save()
                        #c.nivel=nivel
                        #c.seccion= seccion
                        #c.asignatura= asignatura
                        #c.descripcion= descripcion
                        #c.creada_en= creada
                        #c.finaliza_en= finalizada
                        #c.profesor_id=self.request.user.id
                        #c.save()
                        #u = User()
                        #u.username= cedula
                        #u.set_password(cedula)
                        #u.last_name = apellido
                        #u.first_name = nombre
                        #u.dni=cedula
                        #u.email=email
                        #u.save()
                        #Creamos el Estudiante
                        #estudiante = Estudiantes()
                        #estudiante.user = u
                        #estudiante.save()
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        
        date = datetime.today().strftime('%Y-%m-%d  %H:%M:%S')
        fecha_dt = datetime.strptime(date,'%Y-%m-%d  %H:%M:%S')
        print("fecha",fecha_dt)
        id_curso = self.kwargs['pk']
        print("curso",id_curso)
        curso = Cursos.objects.get(id=id_curso)
        info = Mainpage.objects.get(id=1)
        matriculados = Matricula.objects.filter(curso=id_curso)
        asistencias = ListaEstudiantes.objects.filter(curso=id_curso)
        tareas = EntregarTarea.objects.filter(tarea__curso_id=id_curso)
        tareasIndividual = EntregarTarea.objects.filter(tarea__curso=id_curso, estudiante__user__id = self.request.user.id)
        asistenciasIndividuales = ListaEstudiantes.objects.filter(curso=id_curso, estudiant__user__id = self.request.user.id)
        #Quimestre 1
        notaP1 = 0
        notaP2 = 0
        notaEva = 0
        notaPrimerFinal = 0
        primerQuimestreP1 = EntregarTarea.objects.filter(tarea__curso=id_curso, estudiante__user__id = self.request.user.id,tarea__silabo__quimestre__icontains = 'Primero',tarea__silabo__producto__icontains = 'Producto 1')
        primerQuimestreP2 = EntregarTarea.objects.filter(tarea__curso=id_curso, estudiante__user__id = self.request.user.id,tarea__silabo__quimestre__icontains = 'Primero',tarea__silabo__producto__icontains = 'Producto 2')
        primerQuimestreEva = EntregarTarea.objects.filter(tarea__curso=id_curso, estudiante__user__id = self.request.user.id,tarea__silabo__quimestre__icontains = 'Primero',tarea__silabo__producto__icontains = 'Evaluacion')

        #Quimestre 2
        notaP1Q2 = 0
        notaP2Q2 = 0
        notaEvaQ2 = 0
        notaPrimerFinalQ2 = 0
        primerQuimestreP1Q2 = EntregarTarea.objects.filter(tarea__curso=id_curso, estudiante__user__id = self.request.user.id,tarea__silabo__quimestre__icontains = 'Segundo',tarea__silabo__producto__icontains = 'Producto 1')
        primerQuimestreP2Q2 = EntregarTarea.objects.filter(tarea__curso=id_curso, estudiante__user__id = self.request.user.id,tarea__silabo__quimestre__icontains = 'Segundo',tarea__silabo__producto__icontains = 'Producto 2')
        primerQuimestreEvaQ2 = EntregarTarea.objects.filter(tarea__curso=id_curso, estudiante__user__id = self.request.user.id,tarea__silabo__quimestre__icontains = 'Segundo',tarea__silabo__producto__icontains = 'Evaluacion')


        #Quimestre 1
        if len(primerQuimestreP1) > 0:
            acumulador = 0
            for a in primerQuimestreP1:
                acumulador = acumulador + int(a.nota)
                print(a.nota)
            nNotas = len(primerQuimestreP1)
            notaP1 = acumulador / nNotas
        else:
            notaP1 = 0

        if len(primerQuimestreP2) > 0:
            acumuladorP2 = 0
            for a in primerQuimestreP2:
                acumuladorP2 = acumuladorP2 + int(a.nota)
                print(a.nota)
            nNotasp2 = len(primerQuimestreP2)
            notaP2 = acumuladorP2 / nNotasp2
        else:
            notaP2 = 0


        if len(primerQuimestreEva) > 0:
            acumuladorEva = 0
            for a in primerQuimestreEva:
                acumuladorEva = acumuladorEva + int(a.nota)
                print(a.nota)
            nNotasEva = len(primerQuimestreEva)
            notaEva = acumuladorEva / nNotasEva
        else:
            notaEva = 0

        notaPrimerFinal = (((notaP1 + notaP2)/2) * 0.80) + (notaEva * 0.20)



        #Quimestre 2
        if len(primerQuimestreP1Q2) > 0:
            acumuladorQ2 = 0
            for a in primerQuimestreP1Q2:
                acumuladorQ2 = acumuladorQ2 + int(a.nota)
                print(a.nota)
            nNotasQ2 = len(primerQuimestreP1Q2)
            notaP1Q2 = acumuladorQ2 / nNotasQ2
        else:
            notaP1Q2 = 0

        if len(primerQuimestreP2Q2) > 0:
            acumuladorP2Q2 = 0
            for a in primerQuimestreP2Q2:
                acumuladorP2Q2 = acumuladorP2Q2 + int(a.nota)
                print(a.nota)
            nNotasp2Q2 = len(primerQuimestreP2Q2)
            notaP2Q2 = acumuladorP2Q2 / nNotasp2Q2
        else:
            notaP2Q2 = 0


        if len(primerQuimestreEvaQ2) > 0:
            acumuladorEvaQ2 = 0
            for a in primerQuimestreEvaQ2:
                acumuladorEvaQ2 = acumuladorEvaQ2 + int(a.nota)
                print(a.nota)
            nNotasEvaQ2 = len(primerQuimestreEvaQ2)
            notaEvaQ2 = acumuladorEvaQ2 / nNotasEvaQ2
        else:
            notaEvaQ2 = 0

        notaPrimerFinalQ2 = (((notaP1Q2 + notaP2Q2)/2) * 0.80) + (notaEvaQ2 * 0.20)




        print('nota1 ', notaP1)
        print('nota2 ', notaP2)
        print('nota3 ', notaEva)

        for a in tareasIndividual:
            print(a.creada_en)
            if fecha_dt > a.creada_en:
                print("YA VALISTE")
            else:
                print("TODAVIA HAY TIEMPO")

        #print("fecha", date)
        #print("actual", tareasIndividual[0].creada_en)
        #SILABO
        silabo = Silabo.objects.filter(curso=id_curso)
        count = Matricula.objects.filter(curso=id_curso).count
        context['title'] = 'La educación es el camino, no el objetivo.'
        context['curso'] = curso
        context['silabo'] = silabo    
        context['titulo'] = 'Matrículas'
        context['notaP1'] = "{0:.2f}".format(notaP1)
        context['notaP2'] = "{0:.2f}".format(notaP2)
        context['notaEva'] = "{0:.2f}".format(notaEva)
        context['acumulador'] = "{0:.2f}".format(notaPrimerFinal)

        context['notaP1Q2'] = "{0:.2f}".format(notaP1Q2)
        context['notaP2Q2'] = "{0:.2f}".format(notaP2Q2)
        context['notaEvaQ2'] = "{0:.2f}".format(notaEvaQ2)
        context['acumuladorQ2'] = "{0:.2f}".format(notaPrimerFinalQ2)
        context['promedioFinal'] = "{0:.2f}".format(((notaPrimerFinal+notaPrimerFinalQ2)/2))

        context['list_url'] = reverse_lazy('verCurso', kwargs={'pk': self.kwargs['pk']})
        context['action'] = 'add'
        context['info'] = info
        context['date'] = fecha_dt
        context['user'] = self.request.user.username
        context['deberes'] = 'Deberes'
        context['asistenciaList'] = 'Deberes'
        context['post'] = Post.objects.filter(curso=id_curso)
        context['formPost'] = PostForm()
        context['form'] = MatriculaForm()
        context['asistencias'] = asistencias
        context['juegos'] = Juegos.objects.all()
        context['silabos'] = Silabo.objects.filter(curso=id_curso)
        context['matricula'] = matriculados
        context['formAsistencia'] = AsistenciaForm
        context['cursos'] =  id_curso
        context['count'] = count
        context['TareaRecForm'] = TareaRecForm
        context['formCrearTarea'] = CrearTareaForm()
        context['tareas'] = tareas
        context['formFIleTarea'] = TareaFIleForm()
        context['tareasIndividual'] = tareasIndividual
        context['asistenciaIndividual'] = asistenciasIndividuales
        context['silaboForm'] = SilaboForm()
        context['tareaForm'] = TareaForm()
        return context


class MatriculaListView(PermissionMixin, ListView):
    model = Matricula
    template_name = 'matricula/list.html'
    permission_required = 'view_matricula'


    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id_user = self.request.user.id
        print(str(id_user))
        matriculados = Matricula.objects.filter(estudiante__user__id=id_user)
        context['create_url'] = reverse_lazy('cursos_create')
        context['title'] = 'Listado de Matriculas'
        context['object_list'] = matriculados
        return context


def update_nota(request,id):
    nota = request.POST.get('nota')
    estudiante = request.POST.get('id_estudiante')
    print(estudiante)
    print(nota)
    km = EntregarTarea.objects.get(tarea_id = id,estudiante_id = estudiante)
    km.nota = nota
    km.save()
    #messages.success(request, "Estado del pedido actualizado")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))



def update_recuperacion(request,id):
    creada_en = request.POST.get('creada_en')
    estudiante = request.POST.get('id_estudiante')
    print(estudiante)
    print(creada_en)
    km = EntregarTarea.objects.get(tarea_id = id,estudiante_id = estudiante)
    km.creada_en = creada_en
    km.save()
    #messages.success(request, "Estado del pedido actualizado")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def update_silabo(request,id):
    lectivo = request.POST.get('editlectivo')
    actividad = request.POST.get('editActividad')
    descripcion = request.POST.get('editDescripcion')
    link = request.POST.get('editLink')
    id_id = request.POST.get('id_id')

    juego = request.POST.get('id_juego')
    evaluacion = request.FILES.get('evaluacion')
    refuerzo = request.FILES.get('refuerzo')
    archivo = request.FILES.get('archivo')
    silabo = Silabo.objects.get(id = id_id)
    silabo.aLectivo = lectivo
    silabo.actividad = actividad
    silabo.descripcion = descripcion
    silabo.link = link
    if juego:
        silabo.juego_id = juego
    if evaluacion:
        silabo.evaluacion = evaluacion
    if refuerzo:
        silabo.refuerzo = refuerzo
    if archivo:
        silabo.archivo = archivo
    silabo.save()
    #messages.success(request, "Estado del pedido actualizado")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))




def update_file_estudent(request,id):
    archivo = request.FILES.get('archivo')
    estudiante = request.POST.get('id_estudiante')
    print(estudiante)
    km = EntregarTarea.objects.get(tarea_id = id,estudiante_id = estudiante)
    km.archivo = archivo
    km.save()
    #messages.success(request, "Estado del pedido actualizado")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))

def delete_silabo(request,id):
    id_id = id
    silabo = Silabo.objects.get(id = id_id)
    silabo.delete()
    #messages.success(request, "Estado del pedido actualizado")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER'))


class SaleInvoicePdfView(View):

    def get(self, request, *args, **kwargs):
        try:
            template = get_template('matricula/invoiceEstudiant.html')
            date1 = datetime.today().strftime('%Y-%m-%d  %H:%M:%S')
            fecha_dt1 = datetime.strptime(date1,'%Y-%m-%d  %H:%M:%S')
            context = {
                'curso' : Cursos.objects.get(id = self.kwargs['pk']),
                'matricula': Matricula.objects.filter(curso=self.kwargs['pk']),
                'matriculaN': Matricula.objects.filter(curso=self.kwargs['pk']).count,
                'fecha_dt1': fecha_dt1,
            }
            html = template.render(context)
            css_url = os.path.join(settings.BASE_DIR, 'static/lib/bootstrap-4.6.0/css/bootstrap.min.css')
            pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
            return HttpResponse(pdf, content_type='application/pdf')
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('verCurso', kwargs={'pk': self.kwargs['pk']}))


class SaleInvoiceListPdfView(View):

    def get(self, request, *args, **kwargs):
        try:
            template = get_template('matricula/invoiceEstudiantList.html')
            date1 = datetime.today().strftime('%Y-%m-%d  %H:%M:%S')
            fecha_dt1 = datetime.strptime(date1,'%Y-%m-%d  %H:%M:%S')
            context = {
                'curso' : Cursos.objects.get(id = self.kwargs['pk']),
                'lista': ListaEstudiantes.objects.filter(curso=self.kwargs['pk']),
                'listaN': ListaEstudiantes.objects.filter(curso=self.kwargs['pk']).count,
                'fecha_dt1': fecha_dt1,
            }
            html = template.render(context)
            css_url = os.path.join(settings.BASE_DIR, 'static/lib/bootstrap-4.6.0/css/bootstrap.min.css')
            pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
            return HttpResponse(pdf, content_type='application/pdf')
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('verCurso', kwargs={'pk': self.kwargs['pk']}))


class SaleInvoiceNotasPdfView(View):

    def get(self, request, *args, **kwargs):
        try:
            template = get_template('matricula/invoiceEstudiantNotas.html')

            id_curso = self.kwargs['pk']

            #Quimestre 1
            notaP1 = 0
            notaP2 = 0
            notaEva = 0
            notaPrimerFinal = 0
            primerQuimestreP1 = EntregarTarea.objects.filter(tarea__curso=id_curso,tarea__silabo__quimestre__icontains = 'Primero',tarea__silabo__producto__icontains = 'Producto 1')
            primerQuimestreP2 = EntregarTarea.objects.filter(tarea__curso=id_curso,tarea__silabo__quimestre__icontains = 'Primero',tarea__silabo__producto__icontains = 'Producto 2')
            primerQuimestreEva = EntregarTarea.objects.filter(tarea__curso=id_curso,tarea__silabo__quimestre__icontains = 'Primero',tarea__silabo__producto__icontains = 'Evaluacion')

            #Quimestre 2
            notaP1Q2 = 0
            notaP2Q2 = 0
            notaEvaQ2 = 0
            notaPrimerFinalQ2 = 0
            primerQuimestreP1Q2 = EntregarTarea.objects.filter(tarea__curso=id_curso,tarea__silabo__quimestre__icontains = 'Segundo',tarea__silabo__producto__icontains = 'Producto 1')
            primerQuimestreP2Q2 = EntregarTarea.objects.filter(tarea__curso=id_curso,tarea__silabo__quimestre__icontains = 'Segundo',tarea__silabo__producto__icontains = 'Producto 2')
            primerQuimestreEvaQ2 = EntregarTarea.objects.filter(tarea__curso=id_curso,tarea__silabo__quimestre__icontains = 'Segundo',tarea__silabo__producto__icontains = 'Evaluacion')


            #Quimestre 1
            if len(primerQuimestreP1) > 0:
                acumulador = 0
                for a in primerQuimestreP1:
                    acumulador = acumulador + int(a.nota)
                    print(a.nota)
                nNotas = len(primerQuimestreP1)
                notaP1 = acumulador / nNotas
            else:
                notaP1 = 0

            if len(primerQuimestreP2) > 0:
                acumuladorP2 = 0
                for a in primerQuimestreP2:
                    acumuladorP2 = acumuladorP2 + int(a.nota)
                    print(a.nota)
                nNotasp2 = len(primerQuimestreP2)
                notaP2 = acumuladorP2 / nNotasp2
            else:
                notaP2 = 0


            if len(primerQuimestreEva) > 0:
                acumuladorEva = 0
                for a in primerQuimestreEva:
                    acumuladorEva = acumuladorEva + int(a.nota)
                    print(a.nota)
                nNotasEva = len(primerQuimestreEva)
                notaEva = acumuladorEva / nNotasEva
            else:
                notaEva = 0

            notaPrimerFinal = (((notaP1 + notaP2)/2) * 0.80) + (notaEva * 0.20)



            #Quimestre 2
            if len(primerQuimestreP1Q2) > 0:
                acumuladorQ2 = 0
                for a in primerQuimestreP1Q2:
                    acumuladorQ2 = acumuladorQ2 + int(a.nota)
                    print(a.nota)
                nNotasQ2 = len(primerQuimestreP1Q2)
                notaP1Q2 = acumuladorQ2 / nNotasQ2
            else:
                notaP1Q2 = 0

            if len(primerQuimestreP2Q2) > 0:
                acumuladorP2Q2 = 0
                for a in primerQuimestreP2Q2:
                    acumuladorP2Q2 = acumuladorP2Q2 + int(a.nota)
                    print(a.nota)
                nNotasp2Q2 = len(primerQuimestreP2Q2)
                notaP2Q2 = acumuladorP2Q2 / nNotasp2Q2
            else:
                notaP2Q2 = 0


            if len(primerQuimestreEvaQ2) > 0:
                acumuladorEvaQ2 = 0
                for a in primerQuimestreEvaQ2:
                    acumuladorEvaQ2 = acumuladorEvaQ2 + int(a.nota)
                    print(a.nota)
                nNotasEvaQ2 = len(primerQuimestreEvaQ2)
                notaEvaQ2 = acumuladorEvaQ2 / nNotasEvaQ2
            else:
                notaEvaQ2 = 0

            notaPrimerFinalQ2 = (((notaP1Q2 + notaP2Q2)/2) * 0.80) + (notaEvaQ2 * 0.20)
            tareas = EntregarTarea.objects.filter(tarea__curso_id=id_curso)
            date1 = datetime.today().strftime('%Y-%m-%d  %H:%M:%S')
            fecha_dt1 = datetime.strptime(date1,'%Y-%m-%d  %H:%M:%S')
            context = {
                'notaP1' : "{0:.2f}".format(notaP1),
                'notaP2' : "{0:.2f}".format(notaP2),
                'notaEva' : "{0:.2f}".format(notaEva),
                'acumulador' : "{0:.2f}".format(notaPrimerFinal),
                'tareas': tareas,
                'fecha_dt1': fecha_dt1,
                'notaP1Q2' : "{0:.2f}".format(notaP1Q2),
                'notaP2Q2' : "{0:.2f}".format(notaP2Q2),
                'notaEvaQ2' : "{0:.2f}".format(notaEvaQ2),
                'acumuladorQ2' : "{0:.2f}".format(notaPrimerFinalQ2),
                'promedioFinal' : "{0:.2f}".format(((notaPrimerFinal+notaPrimerFinalQ2)/2)),
                'curso' : Cursos.objects.get(id = self.kwargs['pk']),
                'lista': ListaEstudiantes.objects.filter(curso=self.kwargs['pk']),
                'listaN': ListaEstudiantes.objects.filter(curso=self.kwargs['pk']).count,
            }
            html = template.render(context)
            css_url = os.path.join(settings.BASE_DIR, 'static/lib/bootstrap-4.6.0/css/bootstrap.min.css')
            pdf = HTML(string=html, base_url=request.build_absolute_uri()).write_pdf(stylesheets=[CSS(css_url)])
            return HttpResponse(pdf, content_type='application/pdf')
        except:
            pass
        return HttpResponseRedirect(reverse_lazy('verCurso', kwargs={'pk': self.kwargs['pk']}))