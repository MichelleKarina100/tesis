from django.shortcuts import render, redirect
from django.views.generic import CreateView, ListView,UpdateView, DeleteView, FormView, View, TemplateView
from tablib import Dataset

from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from core.users.estudiante.models import Estudiantes
from django.urls import reverse_lazy
from core.users.estudiante.forms import UserForm
from core.user.models import User
from core.users.estudiante.forms import *
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from core.security.mixins import PermissionMixin, ModuleMixin
from django.http import JsonResponse, HttpResponseRedirect, HttpResponse
import json
from core.security.models import *
from core.user.validators import *
from core.users.representate.models import *
import xlrd

from openpyxl import load_workbook
import xlsxwriter
from django.db import transaction


# Create your views here.
class EstudianteListView(PermissionMixin, ListView):
    model = Estudiantes
    template_name = 'estudiante/list.html'
    permission_required = 'view_estudiantes'
    #copiar
    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'upload_excel':
                with transaction.atomic():
                    print("ingresando excel")
                    archive = request.FILES['archive']
                    workbook = load_workbook(filename=archive, data_only=True)
                    excel = workbook[workbook.sheetnames[0]]
                    print("ingresando excel")
                    for row in range(2, excel.max_row + 1):

                        #id = int(excel.cell(row=row, column=1).value)
                        nombre = excel.cell(row=row, column=1).value
                        apellido = excel.cell(row=row, column=2).value
                        cedula = excel.cell(row=row, column=3).value
                        email = excel.cell(row=row, column=4).value
                        email1 = excel.cell(row=row, column=5).value
                        email2 = excel.cell(row=row, column=6).value
                        print('Nombre',nombre)
                        print("Apellido",apellido)
                        print("Cedula",cedula)
                        print("Email",email)
                    
                        u = User()
                        u.username= cedula
                        u.set_password(cedula)
                        u.last_name = apellido
                        u.first_name = nombre
                        u.dni=cedula
                        u.email=email
                        u.save()
                        #Creamos el Estudiante
                        estudiante = Estudiantes()
                        estudiante.user = u
                        estudiante.save()

                        group = Group.objects.get(pk=settings.GROUPS.get('Estudiante'))
                        u.groups.add(group)
            else:
                data['error'] = 'No ha ingresado una opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')
    
    def get_context_data(self, **kwargs):
        print("holas")
        context = super().get_context_data(**kwargs)
        context['create_url'] = reverse_lazy('estudiante_create')
        context['title'] = 'Listado de Estudiantes'
        return context
      
class EstudianteCreateView(PermissionMixin,CreateView):
    model = Estudiantes
    form_class = UserForm
    template_name = 'estudiante/create.html'
    success_url = reverse_lazy('estudiante_list')
    permission_required = 'add_estudiantes'
    

    def validate_data(self):
        data = {'valid': True}
        try:
            type = self.request.POST['type']
            obj = self.request.POST['obj'].strip()
            if type == 'dni':
                if User.objects.filter(dni=obj):
                    data['valid'] = False
            
            elif type == 'email':
                if User.objects.filter(email=obj):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)
    
    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'add':
                contraseña = request.POST['dni']
                form1 = UserForm(request.POST)
                if form1.is_valid():
                    #FORMULARIO PARA EL ESTUDIANTE
                                        #llamo los campos del formulario
                    nombre = request.POST['first_name_repre']
                    apellido = request.POST['last_name_repre']
                    cedula = request.POST['dni_repre']

                    email = request.POST['email_repre']
                    dni = request.POST['dni']

                    valor = vcedula(cedula)
                    valor2 = vcedula(dni)



                    #Formulario del representante
                    if valor == 1 and valor2:
                        form1.instance.username = contraseña
                        form1.save()
                        id_estudiante = form1.instance.id

                        user = User.objects.get(id=id_estudiante)
                        user.set_password(contraseña)
                        user.save()

                        u = Representante()
                        u.nombres_re = nombre
                        u.apellidos_re = apellido
                        u.cedula = cedula
                        u.direccion = cedula
                        u.save()

                        repre = Representante.objects.filter(cedula=cedula).first()

                        #Creamos el Estudiante
                        estudiante = Estudiantes()
                        estudiante.user_id = id_estudiante
                        estudiante.representante_id = repre.id
                        estudiante.save()
                        #Asinamos el grupo correspondiente
                        #GRUPO DE ESTUDIANTE
                        group = Group.objects.get(name="Estudiante")
                        user.groups.add(group)
                    else:
                        messages.error(request, 'Cedulas no validas')


            elif action == 'validate_data':
                return self.validate_data()

            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['list_url'] = self.success_url
        context['title'] = 'Nuevo registro de un Usuario'
        context['action'] = 'add'
        context['form2'] = RepresentanteForm
        return context


class EstudianteUpdateView(PermissionMixin,UpdateView):
    model = Estudiantes
    form_class = UserUpdateForm
    template_name = 'Estudiante/create.html'
    success_url = reverse_lazy('Estudiante:list')
    
    def post(self, request, *args, **kwargs):
        #password = request.POST['password']
        self.object = None
        context = super().get_context_data(**kwargs)
        id_Estudiante = context['pk']
        Estudiante = Estudiantes.objects.get(id=id_Estudiante)
        user = User.objects.get(id=Estudiante.user_id)
        form1 = UserUpdateForm(request.POST, instance=user)
        form2 = EstudianteForm(request.POST, instance=Estudiante, files=request.FILES)
        if form1.is_valid() and form2.is_valid():
            form1.save()
            id_user = form1.instance.id
            user = User.objects.get(id=id_user)
            #user.set_password(password)
            user.save()
            form2.save()
            messages.info(request, 'Datos actualizados correctamente')
            return redirect(self.success_url)
        
        context = self.get_context_data(**kwargs)
        context['form'] = form1
        context['for2'] = form2
        return render(request, self.template_name, context)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        id_user = context['Estudiante'].user_id
        user = User.objects.get(id=id_user)
        Estudiante = Estudiantes.objects.get(user_id=user.id)
        context['title'] = 'Actualizar Técnico'
        context['action'] = 'edit'
        context['entity'] = 'Técnico'
        context['url_imagen'] = Estudiante.imagen
        context['url_firma'] = Estudiante.firma
        context['list_url'] = reverse_lazy('Estudiante:list')
        context['form'] = UserUpdateForm(instance=user)
        context['form2'] = EstudianteForm(instance=Estudiante)
        return context

def activate(request, pk):
    if request.method == 'GET':
        user = User.objects.get(id=pk)
        user.estado = True
        user.save()
        messages.info(request, 'El estado se actualizo correctamente')
        return redirect('Estudiante:list')
    else:
        messages.error(request, 'No se pudo cambiar el estado del Técnico.')
        return redirect('Estudiante:list')

def deactivate(request, pk):
    if request.method == 'GET':
        user = User.objects.get(id=pk)
        user.estado = False
        user.save()
        messages.info(request, 'El estado se actualizo correctamente')
        return redirect('Estudiante:list')
    else:
        messages.error(request, 'No se pudo cambiar el estado del Técnico.')
        return redirect('Estudiante:list')
