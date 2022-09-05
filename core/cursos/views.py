import json

from django.http import JsonResponse, HttpResponse
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView

from core.cursos.forms import *
from core.security.mixins import PermissionMixin
from core.cursos.models import * 
from core.cursos.models import Cursos

from django.contrib import messages
#importar
from openpyxl import load_workbook
from django.db import transaction


class CursosListView(PermissionMixin, ListView):
    model = Cursos
    template_name = 'cursos/list.html'
    permission_required = 'view_cursos'
    
  #copiar
    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        try:
            if action == 'upload_excel':
                with transaction.atomic():
                    archive = request.FILES['archive']
                    workbook = load_workbook(filename=archive, data_only=True)
                    excel = workbook[workbook.sheetnames[0]]
                    for row in range(2, excel.max_row + 1):

                        #id = int(excel.cell(row=row, column=1).value)
                        nivel = excel.cell(row=row, column=1).value
                        seccion = excel.cell(row=row, column=2).value
                        asignatura = excel.cell(row=row, column=3).value
                        descripcion = excel.cell(row=row, column=4).value
                        creada = excel.cell(row=row, column=5).value
                        finalizada = excel.cell(row=row, column=6).value

                        print('Nombre',nivel)
                        print("Apellido",seccion)
                        print("Cedula",asignatura)
                        print("Email",descripcion)
                        print("Email",creada)
                        print("Email",finalizada)
                        print("ass",self.request.user.id)
                        c = Cursos()
                        c.nivel=nivel
                        c.seccion= seccion
                        c.asignatura= asignatura
                        c.descripcion= descripcion
                        c.creada_en= creada
                        c.finaliza_en= finalizada
                        c.profesor_id=self.request.user.id
                        c.save()
                        #u = User()
                        #u.username= cedula
                        #u.set_password(cedula)
                        #u.last_name = apellido
                        #u.first_name = nombre
                        #u.dni=cedula
                        #u.email=email
                        #u.save()
                        #Creamos el Estudiante
                        #estudiante = Estudiantes()
                        #estudiante.user = u
                        #estudiante.save()
            else:
                data['error'] = 'No ha ingresado una opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')
     
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        curos = Cursos.objects.filter(profesor__id = self.request.user.id)
        context['create_url'] = reverse_lazy('cursos_create')
        context['title'] = 'Listado de Cursos'
        context['object_list'] = curos
        return context


class CursosCreateView(PermissionMixin, CreateView):
    model = Cursos
    template_name = 'cursos/create.html'
    form_class = CursoForm
    success_url = reverse_lazy('cursos_list')
    permission_required = 'add_cursos'

    def validate_data(self):
        data = {'valid': True}
        try:
            type = self.request.POST['type']
            obj = self.request.POST['obj'].strip()
            if type == 'name':
                if Cursos.objects.filter(nivel__iexact=obj):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def post(self, request, *args, **kwargs):
        data = {}
        action = request.POST['action']
        form = CursoForm(request.POST)
        try:
            if action == 'add':
                if form.is_valid():
                    form.instance.profesor_id = request.user.id
                    form.save()
                    #messages.info(request, "Grupo de Trabajo Creado Correctamente..")
                    reverse_lazy('cursos_list')
            elif action == 'validate_data':
                return self.validate_data()
            else:
                data['error'] = 'No ha seleccionado ninguna opción'
        except Exception as e:
            data['error'] = str(e)
        return HttpResponse(json.dumps(data), content_type='application/json')

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['list_url'] = self.success_url
        context['title'] = 'Nuevo registro de un Curso'
        context['action'] = 'add'
        return context


class CursosUpdateView(PermissionMixin, UpdateView):
    model = Cursos
    template_name = 'cursos/create.html'
    form_class = CursoForm
    success_url = reverse_lazy('cursos_list')
    permission_required = 'change_cursos'

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def validate_data(self):
        data = {'valid': True}
        try:
            type = self.request.POST['type']
            obj = self.request.POST['obj'].strip()
            id = self.get_object().id

            if type == 'name':
                if Cursos.objects.filter(nivel__iexact=obj).exclude(id=id):
                    data['valid'] = False
        except:
            pass
        return JsonResponse(data)

    def post(self, request, *args, **kwargs):
            data = {}
            action = request.POST['action']
            id_cursos = self.kwargs['pk']
            cursos = Cursos.objects.get(id=id_cursos)
            form = CursoForm(request.POST, instance=cursos, files=request.FILES)
            try:
                if action == 'edit':
                    if form.is_valid():
                        #form.instance.user_creation = request.user.id
                        form.save()
                        #messages.info(request, "Grupo de Trabajo Creado Correctamente..")
                elif action == 'validate_data':
                    return self.validate_data()
                else:
                    data['error'] = 'No ha seleccionado ninguna opción'
            except Exception as e:
                data['error'] = str(e)
            return HttpResponse(json.dumps(data), content_type='application/json')    

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        id_Cursos = self.kwargs['pk']
        curso = Cursos.objects.get(id=id_Cursos)
        context['list_url'] = self.success_url
        context['title'] = 'Edición de un Curso'
        context['form'] = CursoForm(instance=curso)
        context['action'] = 'edit'
        return context


class CursosDeleteView(DeleteView):
    model = Cursos
    template_name = 'cursos/delete.html'
    success_url = reverse_lazy('cursos_list')
    permission_required = 'delete_category'
    url_redirect = success_url

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        data = {}
        try:
            self.object.delete()
        except Exception as e:
            data['error'] = str(e)
        return JsonResponse(data)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Eliminación de un Curso'
        context['entity'] = 'Categorias'
        context['list_url'] = self.success_url
        return context


